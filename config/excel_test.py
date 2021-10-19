from openpyxl import Workbook

#with open('excel_test01','w+',encoding='utf-8') as f:

#test

#获取工作文档
wb = Workbook()
print(wb.active)
wb.create_sheet('123')

print(wb.sheetnames)
#切换工作表
wb = wb['123']
print(wb)