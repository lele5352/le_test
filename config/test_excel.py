import openpyxl
import random


#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('test_excel.xlsx')
#wb.create_sheet('abc')
#wb.create_sheet('123')
ex_list = wb.sheetnames
print(ex_list)

ws = wb['123']
for row in range(1, 5):
    for column in range(1, 5):
        value = random.randint(1, 50)
        ws.cell(row, column, value=value)
    print(row)


#del wb['abc']
ex_list = wb.sheetnames
print(ex_list)

wb.save('test_excel.xlsx')
wb.close()